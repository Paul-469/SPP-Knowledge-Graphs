import openpyxl
import os
import re
# import pandas as pd

os.getcwd()

os.chdir('../resources')


def test(LF):
    #file = 'all-nov-21.xlsx'
    #data = pd.ExcelFile(file)
    #print(data.sheet_names) #this returns the all the sheets in the excel file
    ##['CATALOG']

    #df = data.parse('CATALOG')
    #df.info
    #df.head(10)

    ps = openpyxl.load_workbook('all-nov-21.xlsx')
    sheets = ps.sheetnames
    print(sheets)
    sheet = ps[sheets[0]]    # 'CATALOG'
    print(sheet.max_row)
    search(LF, sheet)


def querryPDC(input):
    ps = openpyxl.load_workbook('all-nov-21.xlsx')
    sheets = ps.sheetnames
    # print(sheets)
    sheet = ps[sheets[0]]  # 'CATALOG'
    # print(sheet.max_row)
    return search(input, sheet)


def search(LF, ws):

    #for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    #    for cell in row:
    #        print(cell)

    # Searching with the names of events as they are in col 2 / min_col=2, max_col=2
    # here the index starts at 1 not 0
    out_list = []
    count = 2
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=ws.max_row, values_only=True):
        if AcronymIsIn(LF, row[0]):
            out_list.append(rowGrap(count, ws))
            # print(row)
            # print(rowGrap(count, ws))
            # print(count)
        count += 1
    return out_list
# Grabbing the entire content of row x as Array. There has to be a better way but it works
def rowGrap(x, ws):
    for row in ws.iter_rows(min_row=x, min_col=1, max_col=17, max_row=x, values_only=True):
        return row

# Trying to prevent hit when the acronym is contained in another acronym; also removing case sensitivity
def AcronymIsIn(x, y):
    y = re.sub(r"[^a-zA-Z0-9]+", ' ', y)
    y = y.lower()
    x = x.lower()
    for word in y.split():
        if x == word:
            return True
